#!/usr/bin/env python
#
# File      : scrape_hackerrank.py
#
# Revision history:
#   Olaf Nielsen / 4.6.2019 / first version. Tested on Windows 10 with python 3.6 and 3.7
#   Olaf Nielsen / 14.6.2019 / hackerranck DOM changed. Additionally, allow for multiple 
#                              solutions per challenge (i.e. different languages)
#   Olaf Nielsen / 17.6.2019 / Added more information to column in the excel output.
#   Olaf Nielsen / 21.6.2019 / json instead of pickle to keep recovery data.
#
# Description:
'''
Extracts code submissions from www.hackerrank.com by scraping the site.

Uses python libraries selenium, BeautifulSoup, openpyxl, json.
selenium uses the chrome browser, and it requires chromedriver.exe
(http://chromedriver.chromium.org/downloads)

Requires environment variables for hackerrank login : HACKERRANK_USER and HACKERRANK_PWD.

Output is written to an Excel file (scrape_hackerrank_<username>.xlsx'). Also, output is 
written to a json-file, which will be used next time the program is run in order to prevent
re-reading pages already known.
'''
# ----------------------------------------------
import re, os, sys, time
import binascii
from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import TimeoutException
from bs4 import BeautifulSoup
import requests
import lxml
from urllib.parse import urljoin
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Color, colors, Border, Side
import pickle
import json

SITE_URL = 'https://www.hackerrank.com'
LOGIN_PAGE = urljoin(SITE_URL, 'auth/login')
FIRST_SUBMISSION_PAGE = urljoin(SITE_URL, 'submissions/all/page/')

# temporary file. No read for re-reading submissions already know from previous runs.
SUBMISSIONS_FILENAME = 'submissions.json'   

# Results are written this Excel file:
OUTPUT_FILENAME      = 'hackerrank_submissions_' + os.getlogin() + '.xlsx'

# ------------------------------------------------------------------------------------
# writeToExcel
# -------------------------------------------------------------------------------------
def writeToExcel(submissions, filename) :
    """
    Writes submissions to an Excel file. Make sure the file is not already opened in Excel.
    """
    def _hyperLinkOf(href, txt) :
        txt = txt.replace('"', "'")
        return f'=HYPERLINK("{challenge_href}", "{txt}")'
    #
    workBook = Workbook()
    excelSheet = workBook.active
    excelSheet.column_dimensions['A'].width = 50
    excelSheet.column_dimensions['B'].width = 13
    excelSheet.column_dimensions['C'].width = 10
    excelSheet.column_dimensions['D'].width = 6
    excelSheet.column_dimensions['E'].width = 10
    excelSheet.column_dimensions['F'].width = 150
    #                     
    excelSheet.append(['Challenge (link - in Excel click \'enable editing\' if not visible)', 'Time', 'Status', 'Points', 'Language', 'Code' ])
    for col in 'ABCDEF' :
        excelSheet[f'{col}1'].border = Border(top=Side(style='medium'), bottom=Side(style='medium'))
        excelSheet[f'{col}1'].alignment = Alignment(wrapText = True, vertical='center')
    #
    excelSheet.title = 'hackerrank submissions'  
    #
    currentRow = 2
    # write each entry in the submissions directory to the excel-sheet. 
    # Items are sorted alphabetically by the text that identifies the challenge (first value in each list submissions[key])
    for (challenge_href, language), \
            (challengeText, timeSubmitted, status, points, _submission_href, code) \
            in sorted(submissions.items(), key=lambda x: x[1][0].upper()):
        # remove this line also to see the errors....
        if status != 'Accepted' : continue
        #
        # column A : link to the challenge 
        excelSheet.cell(row=currentRow, column=1, 
                        value=f'=HYPERLINK("{challenge_href}", "{challengeText}")')
        # setting wrapText=True, makes the entire cell clickable. We don't want that (how to avoid this?)
        excelSheet[f'A{currentRow}'].alignment = Alignment(wrapText=False, vertical='top')
        excelSheet[f'A{currentRow}'].font = Font(bold=True, color=colors.DARKBLUE, underline='single')
        # column B : time submitted
        excelSheet.cell(row=currentRow, column=2, value=timeSubmitted)
        excelSheet[f'B{currentRow}'].alignment = Alignment(vertical='top')
        # column C : status
        excelSheet.cell(row=currentRow, column=3, value=status)
        excelSheet[f'C{currentRow}'].alignment = Alignment(vertical='top')
        # column D : points
        excelSheet.cell(row=currentRow, column=4, value=points)
        excelSheet[f'D{currentRow}'].alignment = Alignment(vertical='top')
        # column E : language
        excelSheet.cell(row=currentRow, column=5, value=language)
        excelSheet[f'E{currentRow}'].alignment = Alignment(vertical='top')
        # column F : code submitted            
        excelSheet.cell(row=currentRow, column=6, value='\n'.join(code).strip())    
        excelSheet[f'F{currentRow}'].alignment = Alignment(wrapText=True, vertical='top')
        excelSheet[f'F{currentRow}'].font = Font(bold=True, name='Courier New')
        excelSheet[f'F{currentRow}'].border = Border(left=Side(style='medium'), right=Side(style='medium'), 
                                                     top=Side(style='medium'), bottom=Side(style='medium'))
        #
        currentRow += 1
    workBook.save(filename) 
# end writeToExcel

# -------------------------------------------------------------------------------------
# find_indexLastPage
# -------------------------------------------------------------------------------------
def find_indexLastPage(driver):
    ''' returns number of submission pages. '''
    max_index = 0
    pagination = driver.find_elements_by_class_name('backbone')
    for element in pagination:
        index = element.get_attribute('data-attr8')
        if index : max_index = max(int(index), max_index)
    return max_index
# end find_indexLastPage

# -------------------------------------------------------------------------------------
# readSubmission
# -------------------------------------------------------------------------------------
def readSubmission(driver, url, count) :
    '''
    Reads the code block in a submission page 
    (e.g. https://www.hackerrank.com/challenges/matrix-rotation-algo/submissions/database/110553576)
    '''
    retry = True; count += 1; timeoutCount = 0
    while retry :
        code = [] 
        try :    
            print (f'{count:>4} opening page {url}')
            driver.get(url)
            # Allow time for loading the page.
            time.sleep(3)                                  
            # Timeout check. If too slow, TimeoutException is raised
            driver.set_page_load_timeout(10)          
            # Check if loading is complete. If not, exception is NoSuchElementException raised.
            driver.find_element_by_class_name('page_footer')
            #
            # Get the list of links ('theme') at the top of the page,
            # e.g. Dashboard > SQL > Aggregation > Population Density Difference.
            # Note : A few pages DON'T have this information, 
            # e.g. https://www.hackerrank.com/challenges/triple-sum/submissions/database/88645926
            theme = driver.find_element_by_css_selector('.content-header > div:nth-child(1) > div:nth-child(1)')
            soup = BeautifulSoup(theme.get_attribute("innerHTML"), 'lxml')
            theme = []
            for link in soup.find_all('a', class_='backbone') :
                href = SITE_URL + link['href']    # we don't really need this....
                theme.append((href, link.text))
            challengeText = ' > '.join([ t[1] for t in theme[1:]]).replace('"', "'")
            #
            # Obtain the code, line by line
            for element in driver.find_elements_by_class_name(' CodeMirror-line ') : 
                code.append(element.text)
            retry = False
        except NoSuchElementException as e:
            timeoutCount += 1
            if timeoutCount == 2 :
                raise e
            print ('Page not ready - retrying...')
        except TimeoutException:
            timeoutCount += 1
            if timeoutCount == 10 :
                raise TimeoutException('Timeout - Something is wrong with the internet connection')
            print (f'Timeout {timeoutCount} - Retrying')
    return [code], challengeText, count
# end readSubmission

# -------------------------------------------------------------------------------------
# saveAlreadyDone
# -------------------------------------------------------------------------------------
def saveAlreadyDone(submissions):
    '''Saves submissions done. Next run will use this file so that only incremental reading is necessary.'''
    with open(SUBMISSIONS_FILENAME, 'w') as file: 
        # json only allows strings as keys. Here the key consists of an url and language, 
        # both not containing '|', so we can do a join on this character.
        json.dump({'|'.join(key): val for key, val in submissions.items()}, file, indent=2)
    # pickle.dump(submissions, open(SUBMISSIONS_FILENAME, 'wb')) 
# end saveAlreadyDone

# -------------------------------------------------------------------------------------
# getAlreadyDone
# -------------------------------------------------------------------------------------
def getAlreadyDone(): # 
    '''returns submissions already done during previous run'''
    alreadyDoneSubmissions = {}
    stillToDoSubmissions = {}
    submissions = {}
    if os.path.exists(SUBMISSIONS_FILENAME) :
        # submissions = pickle.load(open(SUBMISSIONS_FILENAME, 'rb')) 
        with open(SUBMISSIONS_FILENAME, 'r') as file: 
            submissions = json.load(file)   # keys are strings on the form 'url|language'
    #
    # some values may not yet have the code part in them 
    # (e.g. if the program was terminated abnormally in the previous run).
    # We'll just move those to 'stillToDoSubmissions':
    for key, val in submissions.items() :
        if len(val) == 6 :
            alreadyDoneSubmissions[tuple(key.split('|'))] = val
        else :
            stillToDoSubmissions[tuple(key.split('|'))] = val
    return alreadyDoneSubmissions, stillToDoSubmissions
# end getAlreadyDone

# -------------------------------------------------------------------------------------
# getAllSubmissions
# -------------------------------------------------------------------------------------
def getAllSubmissions(driver) :
    '''
    Determines all submissions. Then for each, reads the relevant code data.
    '''
    pageIndex, lastPageIndex = 1, 1
    oldSubmissions, newSubmissions = getAlreadyDone()
    while pageIndex <= lastPageIndex : 
        url = FIRST_SUBMISSION_PAGE + str(pageIndex)
        print (f'opening page {url}')
        driver.get(url)
        # print (driver.page_source)
        # wait for the 'spinner' to complete. There must be a more elegant way to do this ?
        # (google: invisibility_of_element_located)
        time.sleep(3)   
        try :
           submissions = driver.find_element_by_class_name('submissions_list')
           # print (submissions.get_attribute("innerHTML"))
           # check if loading is complete (exception raised, if not)
           driver.find_element_by_class_name('pagination-sub')  
        except NoSuchElementException :
            print ('Retrying...')
            continue
        #
        if pageIndex == 1 : lastPageIndex = find_indexLastPage(driver)
        #
        # switch from selenium to beautifulsoup :
        soup = BeautifulSoup(submissions.get_attribute("innerHTML"), 'lxml')
        # print (soup.prettify())
        # 
        for submission in soup.find_all('div', class_='chronological-submissions-list-view') :
            challenge = submission.find('a', class_='challenge-slug')
            challenge_href = urljoin(SITE_URL, challenge['href'])
            # wait with the challengeText until we get to the code page....
            challengeText = challenge.text  # may be overwritten later....
            language       = submission.find('div', class_='span2 submissions-language').p.text.strip() 
            timeSubmitted  = submission.find('div', class_='span2 submissions-time').p.text.strip()
            status  = submission.find('div', class_='span3').p.text.strip()
            points  = submission.find('div', class_='span1').p.text.strip()
            code_button = submission.find('a', class_='btn btn-inverse view-results backbone')
            submission_href = urljoin(SITE_URL, code_button['href'])
            # note : until 13.6.19, links were on the form
            # https://www.hackerrank.com/challenges/kangaroo/submissions/code/110668369
            # new is, that also
            # https://www.hackerrank.com/challenges/kangaroo/submissions/database/110668369
            # works. Will use this instead because the page has all the code is visible (before only max 25 lines)
            submission_href = submission_href.replace('/code/', '/database/')
            #
            if (challenge_href, language) in oldSubmissions :
                break  # we have the code information already (but timeSubmitted is no longer correct, and there may
                       # have been new submissions to older problems).
                       # Delete SUBMISSIONS_FILENAME to get a full run.
            # if status == 'Accepted' and (challenge_href, language) not in newSubmissions:
            if (challenge_href, language) not in newSubmissions:
                # only one submission per challenge is read. The first encountered is also the latest.
                if status.startswith('Terminated due') : status = 'Timeout'
                newSubmissions[(challenge_href, language)] = [challengeText, timeSubmitted, status, points, submission_href]
        # end for submission in ...
        #
        pageIndex += 1
        if (challenge_href, language) in oldSubmissions : break
    # end while  
    return newSubmissions, oldSubmissions
# end getAllSubmissions
 
# -------------------------------------------------------------------------------------
# site_login
# -------------------------------------------------------------------------------------
def site_login(driver, usr, pwd):
    driver.get(LOGIN_PAGE)
    username = driver.find_element_by_name("username")
    password = driver.find_element_by_name("password")
    username.send_keys(usr)
    password.send_keys(pwd) 
    # find_element_by_class_name should work, but it doesn't ("Compound class names not permitted"):
    driver.find_element_by_xpath("//button[@class='ui-btn ui-btn-large ui-btn-primary auth-button']").click()
    time.sleep(3)
# end site_login

# -------------------------------------------------------------------------------------
# main
# -------------------------------------------------------------------------------------
def main():
    usr = os.getenv('HACKERRANK_USER')
    if not usr:
        print('Define env. var. with Hackerrank user name ($env:HACKERRANK_USER=....)')
        quit()
    pwd = os.getenv('HACKERRANK_PWD')
    if not pwd :
        print ('Define env. var. with Hackerrank password ($env:HACKERRANK_PWD=....)')
        quit()
    #
    # open a browser (uses chromedriver.exe) and login
    driver = webdriver.Chrome()  
    site_login(driver, usr, pwd) 
    #
    # keep this out of the try statement below (because the recover case will not work correctly)
    submissions, submissionsAlreadyDone = getAllSubmissions(driver)
    try :
        count = len(submissionsAlreadyDone)
        print (f'Number of submissions (new + old): {len(submissions)} + {count}')
        for key, val in submissions.items():
            *_, submission_href = val
            submission, challengeText_Temp, count = readSubmission(driver, submission_href, count)
            submissions[key].extend(submission)
            if challengeText_Temp != '' : submissions[key][0] = challengeText_Temp
        # merge new and already done submssions and write them to an excel file:
        submissions.update(submissionsAlreadyDone)
        submissionsAlreadyDone = {}
        writeToExcel(submissions, OUTPUT_FILENAME)
        print ('Result written to ' + OUTPUT_FILENAME)
    finally :
        # for fast recovery, so that pages already read are not repeated after a crash or ctrl-c
        # Note: if doesn't work correctly, if this happens during reading the 
        if submissionsAlreadyDone != {} : submissions.update(submissionsAlreadyDone)
        saveAlreadyDone(submissions)
        # kill the browser
        driver.quit() 
# main

if __name__ == "__main__" :
    try :
        main()
    except KeyboardInterrupt : print ('<ctrl>-c')
    quit()